from openpyxl import load_workbook
import streamlit as st
import pandas as pd
import os
from processador import processar_extrato  # Ajuste conforme o nome correto do seu arquivo ou função
import tempfile


def letra_para_indice(letra):
    letra = letra.upper()
    indice = 0
    for char in letra:
        indice = indice * 26 + (ord(char) - ord('A') + 1)
    return indice - 1

def desmesclar_planilha(arquivo_xlsx):
    try:
        wb = load_workbook(arquivo_xlsx)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.merged_cells.ranges:
                for merged_cell in list(ws.merged_cells):
                    ws.unmerge_cells(range_string=str(merged_cell))
        caminho_desmesclado = arquivo_xlsx.replace(".xlsx", "_desmesclado.xlsx")
        wb.save(caminho_desmesclado)
        return caminho_desmesclado
    except Exception as e:
        st.error(f"Erro ao desmesclar a planilha: {e}")
        return None

def processar_extrato(arquivo_xlsx):
    caminho_desmesclado = desmesclar_planilha(arquivo_xlsx)
    if not caminho_desmesclado:
        return None

    try:
        sheet = pd.read_excel(caminho_desmesclado, sheet_name=0)
    except Exception as e:
        st.error(f"Erro ao abrir a planilha desmesclada: {e}")
        return None

    colunas = {
        'data': 'E', 'valor': 'Y', 'historico': 'G', 'historico_principal': 'G',
        'd_ou_c': 'AE', 'juros': 'AF', 'doc_origem': 'A'
    }

    idx = {k: letra_para_indice(v) for k, v in colunas.items()}
    dados = []
    ultima_data = None

    for _, row in sheet.iterrows():
        data = row.iloc[idx['data']]
        valor = row.iloc[idx['valor']]
        historico = row.iloc[idx['historico']]
        hist_principal = row.iloc[idx['historico_principal']]
        d_ou_c = row.iloc[idx['d_ou_c']]
        juros = row.iloc[idx['juros']]
        doc_origem = row.iloc[idx['doc_origem']]

        if pd.notna(data):
            ultima_data = data
        else:
            data = ultima_data

        dados.append([data, valor, historico, hist_principal, d_ou_c, juros, doc_origem])

    df = pd.DataFrame(dados, columns=[
        "Data", "Valor", "Historico", "Histórico Principal", "D ou C", "Juros/Desconto", "DOC de Origem"
    ])

    df.loc[df['Valor'].notna() & df['Historico'].isna(), 'Historico'] = df['Histórico Principal']

    for i in range(len(df) - 1):
        if pd.notna(df.loc[i, 'Valor']) and pd.notna(df.loc[i, 'Historico']):
            prox_hist = df.loc[i + 1, 'Historico']
            if pd.notna(prox_hist):
                df.loc[i, 'Historico'] = f"{df.loc[i, 'Historico']} - {prox_hist}"

    df = df[df['Valor'].notna()]
    df.drop(columns=['Histórico Principal'], inplace=True)
    df['Tipo de Lançamento'] = df['Data'].apply(lambda x: 'Sintético' if pd.notna(x) else 'Analítico')
    df['Data'] = df['Data'].fillna(method='ffill')

    importar = []
    for i in range(len(df)):
        tipo = df.iloc[i]['Tipo de Lançamento']
        if tipo == 'Analítico':
            importar.append('Sim')
        else:
            if i + 1 < len(df) and df.iloc[i + 1]['Tipo de Lançamento'] == 'Sintético':
                importar.append('Sim')
            else:
                importar.append('Não')
    df['Importar'] = importar

    saida = arquivo_xlsx.replace(".xlsx", "_processado.xlsx")
    df.to_excel(saida, index=False)
    return saida

# ========== INTERFACE STREAMLIT ==========
st.set_page_config(page_title="Processador de Extratos", layout="centered")

st.title("📄 Processador de Extratos Bancários")
st.write("Faça upload de um arquivo `.xlsx` para processar os dados.")

# Upload do arquivo
arquivo = st.file_uploader("Selecione o arquivo Excel (.xlsx)", type=["xlsx"])

# Processamento após upload
if arquivo:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(arquivo.read())
        caminho_temporario = tmp.name

    if st.button("📊 Processar Arquivo"):
        try:
            caminho_saida = processar_extrato(caminho_temporario)

            if caminho_saida:
                with open(caminho_saida, "rb") as f:
                    st.success("✅ Processamento concluído!")
                    st.download_button(
                        label="📥 Baixar Planilha Processada",
                        data=f,
                        file_name=os.path.basename(caminho_saida),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.error("❌ Não foi possível gerar a planilha processada.")

        except Exception as e:
            st.error(f"Erro no processamento: {e}")
